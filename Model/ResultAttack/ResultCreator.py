from audioplayer import AudioPlayer
from openpyxl import load_workbook
import tkinter.ttk as ttk
import os

class ResultCreator(object):

    def __init__(self, attack, resultfile,currentContent):
        self._attack = attack
        self._resultFile = resultfile
        self._paused = False
        self._content = currentContent
        self.load_attack_results()

    def getResultFileExtension(self):
        name_file, file_extension = os.path.splitext(self._resultFile)
        return file_extension

    def load_attack_results(self):
        if self._attack == 'VoIP Eavesdropping':
            if self.getResultFileExtension() == '.wav':
                return self.createAudioPlayer()
        else:
            if self.getResultFileExtension() == '.xlsx':
                return self.parseExcel(self._attack, self._resultFile, self._content)

    def createAudioPlayer(self):
        return AudioPlayer("Model/ResultAttack/Results/VoIP Eavesdropping/" + self._resultFile)

    def parseExcel(self, type_attack, name_file,currentContent):
        filepath = "Model/ResultAttack/Results/" + type_attack + "/" + name_file
        wb = load_workbook(filepath)
        sheet = wb.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        treeview = ttk.Treeview(currentContent, selectmode='browse')
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", "16", "bold"))

        numberColumns = []
        for j in range(1, max_column + 1):
            numberColumns.append(str(j))
        treeview["columns"] = tuple(numberColumns)
        treeview['show'] = 'headings'
        for j in range(1, max_column + 1):
            treeview.column(str(j), anchor='center')
            treeview.heading(str(j), text=sheet.cell(row=1, column=j).value)

        treeview.tag_configure("data", foreground="#C70039", font=("Segoe UI", "12", "bold"))
        for i in range(2, max_row + 1):
            l = []
            for j in range(1, max_column + 1):
                cell_obj = sheet.cell(row=i, column=j)
                l.append(cell_obj.value)
            treeview.insert("", 'end', text="L" + str(i), values=tuple(l), tags=("data",))
        treeview.grid(column=4, row=1, columnspan=max_column, rowspan=max_row, sticky=('EWNS'))
        self.treeview = treeview
        return treeview

    def grid_remove(self):
        if self.treeview != None:
            self.treeview.grid_forget()



