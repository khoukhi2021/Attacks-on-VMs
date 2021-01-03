
import tkinter as Tk
from audioplayer import AudioPlayer
from openpyxl import load_workbook
import tkinter.ttk as ttk
import os

from resultManager.ResultManagerInterface import ResultManagerInterface


class ResultView(object):

    def __init__(self, GUI):
        self._resultManager = ResultManagerInterface()
        self._GUI = GUI
        self._content = GUI.getContent()
        self._resultFile = self._resultManager.create_results(self._GUI.getAttack())
        self._paused = False

    def updateResultFile(self):
        self._resultFile = self._resultManager.create_results(self._GUI.getAttack())

    def getResultFileExtension(self):
        name_file, file_extension = os.path.splitext(self._resultFile)
        return file_extension

    def load_attack_results(self):
        if self._GUI.getAttack() == 'VoIP Eavesdropping':
            if self.getResultFileExtension() == '.wav':
                self.createAudioWidget()
        else:
            if self.getResultFileExtension() == '.xlsx':
                self.parseExcel(self._GUI.getAttack(), self._resultFile)

    def createAudioWidget(self):
        self._GUI.changeMainLabelTitle("RESULTS")
        self._GUI.destroyMainLabel()

        self._resultAudio= AudioPlayer("../View/Resources/Results/" + self._GUI.getAttack() + "/" + self._resultFile)

        playButtonImage = Tk.PhotoImage(file='Resources/Pictures/play.png')
        self._GUI._playButton = Tk.Button(self._content, image=playButtonImage, justify='center',
                                          command=lambda: self._resultAudio.play(block=False))
        self._GUI._playButton.image = playButtonImage

        pauseButtonImage = Tk.PhotoImage(file='Resources/Pictures/pause.png')
        self._GUI._pauseButton = Tk.Button(self._content, image=pauseButtonImage, justify='center',
                                           command=lambda: self.managePause())
        self._GUI._pauseButton.image = pauseButtonImage

        stopButtonImage = Tk.PhotoImage(file='Resources/Pictures/stop.png')
        self._GUI._stopButton = Tk.Button(self._content, image=stopButtonImage, justify='center',
                                          command=lambda: self._resultAudio.close())
        self._GUI._stopButton.image = stopButtonImage

        self._GUI._playButton.grid(column=4, row=3, columnspan=1, rowspan=1, sticky=('EWNS'))
        self._GUI._pauseButton.grid(column=5, row=3, columnspan=1, rowspan=1, sticky=('EWNS'))
        self._GUI._stopButton.grid(column=6, row=3, columnspan=1, rowspan=1, sticky=('EWNS'))

    def parseExcel(self, type_attack, name_file):
        self._GUI.changeMainLabelTitle("RESULTS")
        self._GUI.destroyMainLabel()
        filepath = "../View/Resources/Results/" + type_attack + "/" + name_file
        wb = load_workbook(filepath)
        sheet = wb.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        treeview = ttk.Treeview(self._GUI._content, selectmode='browse')
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", "16", "bold"))

        numberColumns = []
        for j in range(1, max_column + 1):
            numberColumns.append(str(j))
        treeview["columns"] = tuple(numberColumns)
        treeview['show'] = 'headings'
        self._GUI._mainLabel = treeview
        for j in range(1, max_column + 1):
            treeview.column(str(j), anchor='center')
            treeview.heading(str(j), text=sheet.cell(row=1, column=j).value)

        treeview.tag_configure("data", foreground="#C70039", font=("Segoe UI", "12", "bold"))
        for i in range(2, max_row + 1):
            l = []
            for j in range(1, max_column + 1):
                cell_obj = sheet.cell(row=i, column=j)
                l.append(cell_obj.value)
            treeview.insert("", 'end', text="L" + str(i), values=tuple(l), tags=("data",))

        self.treeview_scrollbar = ttk.Scrollbar(self._GUI._content, orient="vertical",
                                                command=self._GUI._mainLabel.yview)
        treeview.configure(yscrollcommand=self.treeview_scrollbar.set)
        self.treeview_scrollbar.grid(row=1, column=9, sticky=('EWNS'))
        treeview.grid(column=4, row=1, columnspan=max_column, rowspan=max_row, sticky=('EWNS'))
        self._GUI.updateMainLabel(treeview)

    def managePause(self):
        if self._paused:
            self.resume()
        else:
            self.pause()

    def pause(self):
        self._paused = True
        self._resultAudio.pause()

    def resume(self):
        self._paused = False
        self._resultAudio.resume()

    def updateResultView(self):
        self.updateResultFile()
        self.load_attack_results()
